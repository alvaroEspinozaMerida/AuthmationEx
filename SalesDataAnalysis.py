import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

"""Creation of Pivot Table"""
df = pd.read_excel('sales_2021.xlsx')

df = df[["Gender", "Product line", "Total"]]

pivot_table = df.pivot_table(index="Gender", columns="Product line", values="Total", aggfunc="sum")

pivot_table.to_excel("pivot_table.xlsx", "Report", startrow= 4)

"""Creation of bar chart on pivot table"""

wb = load_workbook("pivot_table.xlsx")

sheet = wb["Report"]

min_col = wb.active.min_column
max_col = wb.active.max_column

min_row = wb.active.min_row
max_row = wb.active.max_row

for i in range(min_col + 1 , max_col+ 1):
    letter = get_column_letter(i)

    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}2:{letter}3)"
    sheet[f"{letter}{max_row+1}"].style = "Currency"



bar_chart = BarChart()

data = Reference(sheet, min_col=min_col+1,max_col= max_col, min_row=min_row, max_row=max_row)#gets data from the pivot table
categories = Reference(sheet, min_col = min_col, max_col= min_col, min_row=min_row+1, max_row=max_row) # this adds the male and female categories to the chart


bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

bar_chart.title = "Sales by Product line"
bar_chart.style = 3

sheet.add_chart(bar_chart, "B12")
wb.save("barchart.xlsx")

