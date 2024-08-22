from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os
import sys
application_path = os.path.dirname(sys.executable)

month = input("Enter a month: ")

input_path = os.path.join(application_path,'pivot_table.xlsx')

wb = load_workbook(input_path)

sheet = wb["Report"]

min_col = wb.active.min_column
max_col = wb.active.max_column

min_row = wb.active.min_row
max_row = wb.active.max_row

for i in range(min_col + 1 , max_col+ 1):
    letter = get_column_letter(i)

    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}6:{letter}7)"
    sheet[f"{letter}{max_row+1}"].style = "Currency"

sheet["A1"] = "SALES REPORT"
sheet["A2"] = month

sheet["A1"].font = Font("Arial", bold = "True", size = 20)
sheet["A2"].font = Font("Arial", bold = "True", size = 20)

bar_chart = BarChart()

data = Reference(sheet, min_col=min_col+1,max_col= max_col, min_row=min_row, max_row=max_row)#gets data from the pivot table
categories = Reference(sheet, min_col = min_col, max_col= min_col, min_row=min_row+1, max_row=max_row) # this adds the male and female categories to the chart


bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

bar_chart.title = "Sales by Product line"
bar_chart.style = 3

sheet.add_chart(bar_chart, "B12")

output_path = os.path.join(application_path, f'report_{month}.xlsx')
wb.save(output_path)
